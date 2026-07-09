import openpyxl
from pathlib import Path
from typing import List, Dict, Any
from loguru import logger
from ai.document_processing.exceptions import CorruptedDocumentError, EmptyDocumentError


class ExcelReader:
    """Excel document reader using openpyxl.

    Reads multiple sheets, preserves sheet names and row order, and formats tabular data into Markdown.
    """

    def extract_sheets_content(self, file_path: Path) -> List[Dict[str, Any]]:
        """Extracts content from all sheets in the Excel workbook.

        Args:
            file_path: Absolute Path to the XLSX file.

        Returns:
            List[Dict[str, Any]]: List of dictionary elements representing each sheet.
                Each dict has keys 'sheet_name' (str) and 'text' (str in Markdown table format).

        Raises:
            CorruptedDocumentError: If openpyxl fails to load the workbook.
            EmptyDocumentError: If all sheets are empty or workbook has no sheets.
        """
        logger.info("Extracting content from Excel file: {}", file_path.name)
        sheets_content = []

        try:
            # Load the workbook in read-only and data-only modes for fast parsing
            wb = openpyxl.load_workbook(filename=str(file_path), data_only=True, read_only=True)
        except Exception as e:
            logger.error("openpyxl failed to open Excel file {}: {}", file_path.name, str(e))
            raise CorruptedDocumentError(f"Failed to open or parse Excel workbook: {file_path.name}. Error: {str(e)}")

        try:
            sheet_names = wb.sheetnames
            if not sheet_names:
                raise EmptyDocumentError(f"Excel workbook has no sheets: {file_path.name}")

            for name in sheet_names:
                sheet = wb[name]
                sheet_md = self._sheet_to_markdown(sheet)
                if sheet_md.strip():
                    sheets_content.append({"sheet_name": name, "text": sheet_md})

            wb.close()
        except EmptyDocumentError:
            raise
        except Exception as e:
            logger.error("Error occurred while parsing Excel sheets for {}: {}", file_path.name, str(e))
            if 'wb' in locals():
                try:
                    wb.close()
                except Exception:
                    pass
            raise CorruptedDocumentError(f"Failed to process Excel workbook structure: {file_path.name}. Error: {str(e)}")

        if not sheets_content:
            raise EmptyDocumentError(f"No content could be extracted from Excel workbook: {file_path.name}")

        return sheets_content

    def _sheet_to_markdown(self, sheet) -> str:
        """Converts an openpyxl sheet to a markdown table structure.

        Args:
            sheet: The openpyxl Worksheet object.

        Returns:
            str: Markdown table formatted representation of sheet rows.
        """
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return ""

        # Clean rows by dropping trailing empty rows
        cleaned_rows = []
        for r in rows:
            # Check if the row contains at least one cell with non-empty text/value
            if any(cell_val is not None and str(cell_val).strip() != "" for cell_val in r):
                cleaned_rows.append(r)

        if not cleaned_rows:
            return ""

        markdown_lines = []
        for r_idx, row in enumerate(cleaned_rows):
            row_vals = []
            for cell in row:
                if cell is None:
                    row_vals.append("")
                else:
                    # Clean newlines inside the cells to keep rows formatted cleanly as Markdown lines
                    row_vals.append(str(cell).strip().replace("\n", " "))

            markdown_lines.append("| " + " | ".join(row_vals) + " |")

            # Append the standard Markdown column division line after the header (row 1)
            if r_idx == 0 and len(cleaned_rows) > 1:
                separators = ["---"] * len(row_vals)
                markdown_lines.append("| " + " | ".join(separators) + " |")

        return "\n".join(markdown_lines)
